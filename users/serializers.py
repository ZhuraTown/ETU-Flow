from django.http import QueryDict
from rest_framework import serializers
from rest_framework.exceptions import NotFound
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework.validators import UniqueValidator

from openpyxl import load_workbook, Workbook

from etu_flow_api.settings import WORK_SHEET
from users.models import UserETU


class UserETUSerializer(serializers.ModelSerializer):
    password = serializers.CharField(write_only=True, required=True)
    email = serializers.EmailField(required=True,
                                   validators=[UniqueValidator(queryset=UserETU.objects.all())])
    last_name = serializers.CharField(required=True)
    first_name = serializers.CharField(required=True)
    is_staff = serializers.BooleanField(required=False)

    class Meta:
        model = UserETU
        fields = (
            "id",
            "email",
            "password",
            "user_type",
            "first_name",
            "last_name",
            "middle_name",
            "coin_balance",
            "profile_image",
            "phone_number",
            "group",
            "is_staff")
        read_only_fields = ("id", )

    def validate(self, attrs):
        if not self.context.get('request').user.is_staff and attrs.get('is_staff'):
            raise serializers.ValidationError("У вас недостаточно прав, на приглашение администратора")
        return attrs

    def create(self, validated_data):
        user = UserETU.objects.create(**validated_data)
        password = self.context.get('request').data.get('password')
        user.set_password(password)
        user.save()
        return user


class UserETUUpdateSerializer(serializers.ModelSerializer):
    first_name = serializers.CharField(required=False)
    last_name = serializers.CharField(required=False)
    middle_name = serializers.CharField(required=False)
    phone_number = serializers.CharField(required=False)

    class Meta:
        model = UserETU
        fields = (
            "id",
            "email",
            "user_type",
            "first_name",
            "last_name",
            "middle_name",
            "coin_balance",
            "profile_image",
            "phone_number",
            "is_staff",
            "group")
        read_only_fields = ("id", "email", "user_type", "coin_balance", "group", "is_staff")


class UserETUUpdateAdminSerializer(serializers.ModelSerializer):

    class Meta:
        model = UserETU
        fields = (
            "id",
            "email",
            "user_type",
            "first_name",
            "last_name",
            "middle_name",
            "coin_balance",
            "profile_image",
            "phone_number",
            "is_staff",
            "group")
        read_only_fields = ("id", "email", "is_staff")


class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    def validate(self, attrs):
        data = {}
        user = UserETU.objects.filter(email=attrs['email']).first()
        if not user or not user.check_password(attrs['password']):
            raise NotFound(detail='Неправильный логин или пароль')
        refresh = self.get_token(user)
        data['refresh'] = str(refresh)
        data['access'] = str(refresh.access_token)
        data['user'] = UserETUSerializer(instance=user).data
        return data


class StudentsUploadSerializer(serializers.Serializer):
    file = serializers.FileField(write_only=True, required=True)

    def to_internal_value(self, data: QueryDict):
        file = data.pop("file")
        wb = load_workbook(file[0])
        data['wb'] = wb  # type:
        ws = wb.get_sheet_by_name(WORK_SHEET)
        if ws:
            data['ws'] = ws
            data['table'] = ws[ws.dimensions]
        return data

    def validate(self, attrs: QueryDict):
        wb = attrs.get('wb')

        if not isinstance(wb, Workbook):
            raise serializers.ValidationError('Неправильный формат файла')
        if not wb.get_sheet_by_name(WORK_SHEET):
            raise serializers.ValidationError("Рабочий лист <Сотрудники> не найден")
        return attrs

    def create(self, validated_data):
        table = validated_data.pop('table')
        group = validated_data.pop('group')
        students = UserETU.create_from_xlsx(group, table)
        return students

    def to_representation(self, instance):
        response = super().to_representation(instance)
        response['students'] = [UserETUSerializer(instance=student).data for student in instance]
        return response

